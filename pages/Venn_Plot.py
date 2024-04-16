import pandas as pd
import streamlit as st
import os
import base64
from io import BytesIO
from openpyxl import Workbook
from venn import venn
import matplotlib.pyplot as plt


st.set_page_config(layout="wide")

# URL of the image from the web
image_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcREh-y7VJtrA03RIlxLNVxt0DUOZyGBXELj1vqaAm_c1kWOW0RUqdP7QrysLqvZ2tSLUVj6acdWlUI&usqp=CAU&ec=48665698"

# Display the image using Streamlit's image function
st.write("<p style='text-align:right;'><img src='"+image_url+"' width=250 height=175></p>",unsafe_allow_html=True)

# Web App Title
st.title("Venn_Plot")

st.subheader("Upload your files")


# Function to create dataframes for all uploaded files
def create_dataframes(uploaded_files):
    dataframes = []
    for input_file in uploaded_files:
        try:
            # Load the dataframe from the input file
            df = pd.read_excel(input_file)
            
            # Use a unique identifier for dataframe names without the file extension
            file_name_without_extension = input_file.name.split('.')[0]
            df.name = f"{file_name_without_extension}"
            
            # Check if file already exists
            if any(existing_df.name == df.name for existing_df in dataframes):
                st.warning(f"{df.name} already exists. Please choose a different file name.")
            else:
                dataframes.append(df)
                st.success(f"{input_file.name} uploaded successfully!")
        except Exception as e:
            st.error(f"Error reading {input_file.name}: {e}")

    return dataframes


# Function to select a single column from all dataframes
def select_column_from_dataframes(dataframes, column_name):
    selected_columns = {}
    for df in dataframes:
        if column_name.lower() in df.columns.str.lower():
            selected_columns[df.name] = set(df[column_name])  # Convert column to set
        else:
            st.warning(f"Column '{column_name}' not found in '{df.name}' dataframe.")
    return selected_columns

def calculate_set_operations(file_names, *sets):
    intersections = {}
    uniques = {}
    common_elements = sets[0].intersection(*sets[1:])

    # Calculate intersections
    for i, set1 in enumerate(sets):
        for j, set2 in enumerate(sets[i + 1:], i + 1):
            intersection_key = f"Intersection between {file_names[i]} and {file_names[j]}"
            intersection = set1.intersection(set2)
            intersections[intersection_key] = intersection
            common_elements = common_elements.intersection(intersection)

            # Calculate intersections for three sets
            for k, set3 in enumerate(sets[j + 1:], j + 1):
                intersection_key = f"Intersection between {file_names[i]}, {file_names[j]}, and {file_names[k]}"
                intersection = intersection.intersection(set3)
                intersections[intersection_key] = intersection
                common_elements = common_elements.intersection(intersection)

                # Calculate intersections for four sets
                for l, set4 in enumerate(sets[k + 1:], k + 1):
                    intersection_key = f"Intersection between {file_names[i]}, {file_names[j]}, {file_names[k]}, and {file_names[l]}"
                    intersection = intersection.intersection(set4)
                    intersections[intersection_key] = intersection
                    common_elements = common_elements.intersection(intersection)

                    # Calculate intersections for five sets
                    for m, set5 in enumerate(sets[l + 1:], l + 1):
                        intersection_key = f"Intersection between {file_names[i]}, {file_names[j]}, {file_names[k]}, {file_names[l]}, and {file_names[m]}"
                        intersection = intersection.intersection(set5)
                        intersections[intersection_key] = intersection
                        common_elements = common_elements.intersection(intersection)

    # Calculate unique values
    for i, s in enumerate(sets):
        unique_key = f"Unique to {file_names[i]}"
        uniques[unique_key] = s.difference(*[sets[j] for j in range(len(sets)) if j != i])

    # Add common elements key
    intersections["Common Elements"] = common_elements
    return intersections, uniques

# Function to download set operations results as an Excel file
def download_set_operations_as_excel(intersections, uniques):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Set Operations'

    # Write intersections and unique values in the same column
    col_num = 1
    for key, values in intersections.items():
        ws.cell(row=1, column=col_num, value=f'{key} (Intersection)')
        for row_num, value in enumerate(values, start=2):
            ws.cell(row=row_num, column=col_num, value=value)
        col_num += 1

    for key, values in uniques.items():
        ws.cell(row=1, column=col_num, value=f'{key} (Unique)')
        for row_num, value in enumerate(values, start=2):
            ws.cell(row=row_num, column=col_num, value=value)
        col_num += 1

    # Save the workbook to BytesIO
    output_file = BytesIO()
    wb.save(output_file)
    output_file.seek(0)

    return output_file
    
# Main Streamlit app code
uploaded_files = st.file_uploader(
    label="hi",
    type="xlsx",
    label_visibility="collapsed",
    accept_multiple_files=True
)
st.write("User can input multiple files but the file should have same column name in all the file")
st.write("[Sample-Input](https://docs.google.com/spreadsheets/d/1dVtTVpuDgVLeVv4lgvoYH8zMc5Y5AfaQ/edit?usp=sharing&ouid=103232618408666892680&rtpof=true&sd=true)")
st.write("[Sample-Output](https://docs.google.com/spreadsheets/d/1K0HRInvjFL2_aiK9jRoKTa8Dq43kfFtx/edit?usp=sharing&ouid=103232618408666892680&rtpof=true&sd=true)")

if uploaded_files is not None:
    # Call the create_dataframes function to create dataframes for all uploaded files
    dataframes = create_dataframes(uploaded_files)
    st.write(f"Created {len(dataframes)} dataframes from uploaded files:")
    
    # Get the column name from the user (you can modify this part based on how the user provides the column name)
    column_name = st.text_input("Enter the column name to select:")
    
    if column_name:
    # Call the select_column_from_dataframes function to select the desired column from all dataframes
        selected_columns = select_column_from_dataframes(dataframes, column_name)

        if selected_columns:
            st.write("Selected columns:")
            for name, column_data in selected_columns.items():
                st.write(f"DataFrame: {name}")

            # Create a dictionary of sets for all selected columns
            all_selected_columns = {name: set(column_data) for name, column_data in selected_columns.items()}
            file_names = list(selected_columns.keys())  # Convert keys to a list
            sets = list(selected_columns.values())       # Convert values to a list
            
            st.title("Venn Diagram")
            font_size = st.slider("Select font size:", 6, 16, 10)
            # Display the venn diagram
            fig, ax = plt.subplots(figsize=(12,12))
            venn(all_selected_columns, ax=ax, legend_loc="upper left", fontsize=font_size)
            ax.set_title("Venn Diagram", fontsize=font_size+4)
            st.pyplot(fig)

            intersections, uniques = calculate_set_operations(file_names, *sets)

           # Add a download button to download the set operations results as an Excel file
            download_button = st.button("Download Set Operations Results as Excel")
            if download_button:
                output_file = download_set_operations_as_excel(intersections, uniques)
                st.download_button(
                    label="Click here to download",
                    data=output_file,
                    file_name="set_operations_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        else:
            st.warning(f"No data found for column '{column_name}'.")
