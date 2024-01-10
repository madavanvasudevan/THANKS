import pandas as pd
import streamlit as st
from venn import venn
from openpyxl import Workbook
from io import BytesIO
import matplotlib.pyplot as plt
import os

st.set_page_config(layout="wide")

# URL of the image from the web
image_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcREh-y7VJtrA03RIlxLNVxt0DUOZyGBXELj1vqaAm_c1kWOW0RUqdP7QrysLqvZ2tSLUVj6acdWlUI&usqp=CAU&ec=48665698"

# Display the image using Streamlit's image function
st.write("<p style='text-align:right;'><img src='"+image_url+"' width=250 height=150></p>",unsafe_allow_html=True)

# Web App Title
st.markdown('''
# **VennGenius**
''')

st.subheader("Upload your files")

# Function to create dataframes for all uploaded files
def create_dataframes(uploaded_files):
    dataframes = {}
    for input_file in uploaded_files:
        # Check if file already exists
        if input_file.name in dataframes:
            st.warning(f"{input_file.name} already exists. Please choose a different file name.")
        else:
            # Load the dataframe from the input file
            df = pd.read_excel(input_file)
            
            # Check if dataframe has 3 columns
            if len(df.columns) != 3:
                st.warning(f"{input_file.name} should have 3 columns.")
            # Check if the first column is text and the other two are numeric
            elif not (df.iloc[:,0].dtype == 'object' and 
                      df.iloc[:,1].dtype == 'float64' and 
                      df.iloc[:,2].dtype == 'float64'):
                st.warning(f"{input_file.name} should have the first column as text and the other two as numeric.")
            else:
                # Do something with the dataframe
                st.success(f"{input_file.name} uploaded successfully!")
                
                # Use file name as dataframe name
                dataframes[input_file.name] = df
    return dataframes

def calculate_set_operations(file_names, *sets):
    intersections = {}
    uniques = {}
    common_elements = sets[0].intersection(*sets[1:])

    # Calculate intersections
    for i, set1 in enumerate(sets):
        for j, set2 in enumerate(sets[i + 1:], i + 1):
            intersection_key = f"Intersection between {file_names[i]} and {file_names[j]}"
            intersection = set1.intersection(set2)
            intersections[intersection_key] = intersection
            common_elements = common_elements.intersection(intersection)

            # Calculate intersections for three sets
            for k, set3 in enumerate(sets[j + 1:], j + 1):
                intersection_key = f"Intersection between {file_names[i]}, {file_names[j]}, and {file_names[k]}"
                intersection = intersection.intersection(set3)
                intersections[intersection_key] = intersection
                common_elements = common_elements.intersection(intersection)

                # Calculate intersections for four sets
                for l, set4 in enumerate(sets[k + 1:], k + 1):
                    intersection_key = f"Intersection between {file_names[i]}, {file_names[j]}, {file_names[k]}, and {file_names[l]}"
                    intersection = intersection.intersection(set4)
                    intersections[intersection_key] = intersection
                    common_elements = common_elements.intersection(intersection)

                    # Calculate intersections for five sets
                    for m, set5 in enumerate(sets[l + 1:], l + 1):
                        intersection_key = f"Intersection between {file_names[i]}, {file_names[j]}, {file_names[k]}, {file_names[l]}, and {file_names[m]}"
                        intersection = intersection.intersection(set5)
                        intersections[intersection_key] = intersection
                        common_elements = common_elements.intersection(intersection)

    # Calculate unique values
    for i, s in enumerate(sets):
        unique_key = f"Unique to {file_names[i]}"
        uniques[unique_key] = s.difference(*[sets[j] for j in range(len(sets)) if j != i])

    # Add common elements key
    intersections["Common Elements"] = common_elements
    return intersections, uniques
    
# Function to download set operations results as an Excel file
def download_set_operations_as_excel(intersections, uniques):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Set Operations'

    # Write intersections
    row_num = 1
    for key, values in intersections.items():
        ws.cell(row=row_num, column=1, value=key)
        for col_num, value in enumerate(values, start=2):
            ws.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    # Write unique values
    offset = len(intersections) + 2  # Add some space between intersections and unique values
    for key, values in uniques.items():
        ws.cell(row=offset, column=1, value=key)
        for col_num, value in enumerate(values, start=2):
            ws.cell(row=offset, column=col_num, value=value)
        offset += 1

    # Save the workbook to BytesIO
    output_file = BytesIO()
    wb.save(output_file)
    output_file.seek(0)
    
    return output_file
    
# Main Streamlit app code
uploaded_files = st.file_uploader(
    label="hi",
    type="xlsx",
    label_visibility="collapsed",
    accept_multiple_files=True,
    key="VennGenius"
)
st.write("User can input up-to 6 files but the file should be in the format of sample input")
st.write("[Sample-Input](https://docs.google.com/spreadsheets/d/1BJcJ2djLZTbQv-yA0UFJHbsIvZclc8FW/edit?usp=share_link&ouid=103232618408666892680&rtpof=true&sd=true)")
st.write("[Sample-Output](https://drive.google.com/file/d/1WG5--ruEnO44GAUgvwlhEyDM0S8tgFUY/view?usp=share_link)")

if uploaded_files is not None:
    # Call the create_dataframes function to create dataframes for all uploaded files
    dataframes = create_dataframes(uploaded_files)
    st.write(f"Created {len(dataframes)} dataframes from uploaded files:")
    
    # Create dictionaries to store the up and down dataframes for each file
    up_dfs = {}
    down_dfs = {}
    
    # Ask the user for input
    logFC_upregulator = st.number_input("Select the uplogFC threshold value:", key='logFC_upinput', min_value=0.0, step=0.1)
    logFC_downregulator = st.number_input("Select the downlogFC threshold value:", key='logFC_downinput', min_value=-1e6, max_value=-0.1, step=-0.1, value=-0.1)
    
    # Add a submit button
    submit_button = st.checkbox(label='Submit')
    if submit_button:
        for df_name, df in dataframes.items():
            # Remove file extension from dataframe name
            df_name_clean = df_name.split('.')[0]
            
            # Create two sub-dataframes based on the values in column 'logFC'
            df_up = df[df['logFC'] >= logFC_upregulator]
            df_down = df[df['logFC'] <= logFC_downregulator]    
            # Store the sub-dataframes in the dictionaries
            up_dfs[df_name_clean] = df_up
            down_dfs[df_name_clean] = df_down
    
    # Convert sub-dataframes to sets
    if len(up_dfs) > 0 and len(down_dfs) > 0:
        up_sets = {name: set(df.iloc[:, 0].tolist()) for name, df in up_dfs.items()}
        down_sets = {name: set(df.iloc[:, 0].tolist()) for name, df in down_dfs.items()}
        
        
        # Create a Streamlit app
        st.title("Venn Diagram")
        font_size = st.slider("Select font size:", 6, 16, 10)
        # Display the venn diagram
        fig, ax = plt.subplots(figsize=(12,12))
        venn(up_sets, ax=ax, legend_loc="upper left", fontsize = font_size)
        ax.set_title("UpRegulated Venn Diagram", fontsize=font_size+4)
        st.pyplot(fig)
        # Set operations for 'up_sets'
        file_names1 = list(up_sets.keys())
        sets1 = list(up_sets.values())
        intersections1, uniques1 = calculate_set_operations(file_names1, *sets1)

        # Add a download button for 'up_sets'
        download_button_up = st.button("Download Up Set Operations Results as Excel")
        if download_button_up:
            output_file_up = download_set_operations_as_excel(intersections1, uniques1)
            st.download_button(
                label="Click here to download",
                data=output_file_up,
                file_name="Up_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Display the venn diagram
        fig1, ax = plt.subplots(figsize=(12,12))
        venn(down_sets, ax=ax, legend_loc="upper left", fontsize = font_size)
        ax.set_title("DownRegulated Venn Diagram", fontsize=font_size+4)
        st.pyplot(fig1)
        # Set operations for 'down_sets'
        file_names2 = list(down_sets.keys())
        sets2 = list(down_sets.values())
        intersections2, uniques2 = calculate_set_operations(file_names2, *sets2)

        # Add a download button for 'down_sets'
        download_button_down = st.button("Download Down Set Operations Results as Excel")
        if download_button_down:
            output_file_down = download_set_operations_as_excel(intersections2, uniques2)
            st.download_button(
                label="Click here to download",
                data=output_file_down,
                file_name="Down_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
